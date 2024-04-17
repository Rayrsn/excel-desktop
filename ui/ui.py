# This Python file uses the following encoding: utf-8

# Important:
# You need to run the following command to generate the ui_form.py file
#     pyside6-uic form.ui -o ui_form.py, or
#     pyside2-uic form.ui -o ui_form.py

import sys, os

from PySide6.QtWidgets import (
    QApplication,
    QMainWindow,
    QTableWidgetItem,
    QWidget,
    QFileDialog,
    QMessageBox,
    QVBoxLayout,
    QHBoxLayout,
    QTableWidget,
    QLabel,
    QPushButton,
    QLineEdit,
    QDialog,
    QComboBox,
    QGridLayout,
    QSpacerItem,
    QProgressBar,
)

from PySide6.QtGui import QPixmap, QFont, QIcon
from PySide6.QtCore import Qt, QThread, Signal

from ui.ui_form import Ui_MainWindow
from update_doc_file import gen_docs

import openpyxl
from utils.btn import (
    generate_monthly_cases_report,
    generate_weekly_cases_report,
    generate_legal_aid_report,
    generate_bail_refused_report,
    generate_empty_counsel_report,
    generate_non_zero_balance_report,
    generate_stage_reports,
)

import ui.network as network
import os

if os.environ.get("SERVER_URL"):
    URL = os.environ.get("SERVER_URL")
else:
    URL = "https://excel-api.fly.dev"
DATA = {}
LAST_PAGE = 0

class MainWindow(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        
        win_icon = QIcon()
        if hasattr(sys, "_MEIPASS"):
            win_icon.addFile(sys._MEIPASS + "/bkp_logo.ico")
        else:
            win_icon.addFile('bkp_logo.ico')
        self.setWindowIcon(win_icon)
        
        self.ui.exitbutton.clicked.connect(self.closeApplication)
        # DISABLE IMPORT BUTTON ###################
        # self.ui.importbutton.clicked.connect(self.openFile)
        self.ui.refreshbutton.clicked.connect(lambda: self.loadJsonData(URL, last_page=True))
        self.ui.exportbutton.clicked.connect(lambda: self.genDocsBtn(DATA))
        self.ui.newentrybutton.clicked.connect(self.showNewEntryDialog)
        self.ui.deleteentrybutton.clicked.connect(self.showDeleteEntryDialog)
        self.ui.operationsbutton.clicked.connect(self.showOperationDialog)
        self.ui.exitbutton.setCursor(Qt.PointingHandCursor)
        # DISABLE IMPORT BUTTON ###################
        # self.ui.importbutton.setCursor(Qt.PointingHandCursor)
        self.ui.exportbutton.setCursor(Qt.PointingHandCursor)
        self.ui.newentrybutton.setCursor(Qt.PointingHandCursor)
        self.ui.deleteentrybutton.setCursor(Qt.PointingHandCursor)
        self.ui.refreshbutton.setCursor(Qt.PointingHandCursor)
        self.ui.operationsbutton.setCursor(Qt.PointingHandCursor)
        self.first_run_entry = True

        self.setStyleSheet(
            """
            QMainWindow {
                background-color: #333;
            }

            QTableWidget {
                gridline-color: #999;
                font-size: 14px;
            }

            QTableWidget QHeaderView::section {
                background-color: #666;
                color: #fff;
                padding: 5px;
                border: 1px solid #999;
            }

            QTableWidget QTableCornerButton::section {
                background-color: #666;
                border: 1px solid #999;
            }

            QWidget > QPushButton {
                background-color: #007bff; /* Green */
                border: none;
                color: white;
                padding: 15px 32px;
                text-align: center;
                text-decoration: none;
                font-size: 16px;
                margin: 4px 2px;
            }

            QWidget > QPushButton:hover {
                background-color: #3094fd;
            }
        """
        )
        # DISABLE IMPORT BUTTON ###################
        self.ui.importbutton.setStyleSheet("background-color: gray;") 

        # clear existing tabs
        self.ui.tabWidget.clear()

        # create a new tab for the logo
        tab = QWidget()
        tab.setObjectName("logoTab")
        self.ui.tabWidget.addTab(tab, "Start")

        # create a QLabel for the logo
        logoLabel = QLabel(tab)
        if hasattr(sys, "_MEIPASS"):
            logoPixmap = QPixmap(sys._MEIPASS + "/bkp_logo.jpg")
        else:
            logoPixmap = QPixmap("bkp_logo.jpg")
        logoLabel.setPixmap(logoPixmap)
        textLabel = QLabel("BKP Solicitors Client Data", tab)
        font = QFont("Calibri", 72)
        textLabel.setFont(font)

        # create a layout for the tab and add the logoLabel
        hboxLayout = QHBoxLayout()
        hboxLayout.addWidget(logoLabel)
        hboxLayout.addWidget(textLabel)
        tab.setLayout(hboxLayout)

    def loadJsonDataFinished(self, data):
        if data == "Error" or data is None:
            self.showAlarm("Network error", "Failed to fetch data from the server!")
            # close loading dialog
            self.loading_dialog.close()
            return
        global DATA
        DATA = data
        json_data = data
        
        sheets = network.get_sheets(json_data)

        # close loading dialog
        self.loading_dialog.close()
        
        # clear existing tabs
        self.ui.tabWidget.clear()

        # create tabs
        for _ in range(len(sheets)):
            tab = QWidget()
            tab.setObjectName("tab")
            self.ui.tabWidget.addTab(tab, "")

        # show data of JSON in qt table
        for sh_num, sheet in enumerate(sheets):
            # Create a new QTableWidget for this tab
            tableWidget = QTableWidget()
            tableWidget.setRowCount(network.get_row_count(json_data, sheet))
            tableWidget.setColumnCount(network.get_column_count(json_data, sheet))
            # Enable sorting
            tableWidget.setSortingEnabled(True)
            # set default sorting by first column
            # tableWidget.sortItems(0)

            # Add the QtableWidget to a QHBoxLayout inside a QVBoxLayout
            hboxLayout = QHBoxLayout()
            hboxLayout.addWidget(tableWidget)
            vboxLayout = QVBoxLayout()
            vboxLayout.addLayout(hboxLayout)
            self.ui.tabWidget.widget(sh_num).setLayout(vboxLayout)

            # Change tabs name
            self.ui.tabWidget.setTabText(sh_num, sheet)

            # set value of table from JSON data
            headers = network.get_headers(json_data, sheet)
            if "id" in headers:
                headers.remove("id")
            tableWidget.setHorizontalHeaderLabels(headers)
            
            # sort the data in each sheet by the Sr_No column

            for i in range(tableWidget.rowCount()):
                for j in range(tableWidget.columnCount()):
                    headers = list(headers)
                    cell_data = network.get_data_from_cell(json_data, sheet, i, headers[j])
                    # if cell_data is float then convert it to int
                    if cell_data == "__null__":
                        cell_data = ""
                    if cell_data is not None and cell_data != "" and cell_data != "__null__":
                        if isinstance(cell_data, float) or (isinstance(cell_data, str) and cell_data.replace('.', '', 1).isdigit()):
                            cell_data = int(float(cell_data))
                        # if in column sr no, then convert it to int
                        if headers[j] == "Sr_No":
                            try:
                                cell_data = int(cell_data)
                            except Exception as e:
                                print(f"Error in row {i} and column {j} and table {sheet}, cell data: {cell_data}: {e}")
                    # Skip setting the item if the header is "id"
                    if headers[j] != "id":
                        tableWidget.setItem(i, j, QTableWidgetItem(str(cell_data)))
            
            # resize columns to fit the contents
            tableWidget.resizeColumnsToContents()
            # on cell change print the changed cell
            tableWidget.cellChanged.connect(self.handleCellChanged)

    def loadJsonData(self, url, last_page=False):
        """
        load JSON data into qt tables
        """
        if last_page:
            global LAST_PAGE
            # get the current page to the last page
            LAST_PAGE = self.ui.tabWidget.currentIndex()
        # make a loading dialog
        self.loading_dialog = LoadingDialog(self)
    
        self.fetchDataThread = FetchDataThread(url)
        self.fetchDataThread.dataReady.connect(self.loadJsonDataFinished)
        self.fetchDataThread.start()
        
        self.loading_dialog.exec()
        
        # switch to the last page
        if last_page:
            self.ui.tabWidget.setCurrentIndex(LAST_PAGE)

    def addSheetToTabs(self, sheet_name, data):
        # Create a new tab
        tab = QWidget()
        tab.setObjectName("tab")
        self.ui.tabWidget.addTab(tab, "")
        # Create a new QTableWidget for this tab
        tableWidget = QTableWidget()
        tableWidget.setRowCount(len(data[next(iter(data))]))
        tableWidget.setColumnCount(len(data))
        # Enable sorting
        tableWidget.setSortingEnabled(True)
        # set default sorting by first column
        tableWidget.sortItems(0)

        # Add the QtableWidget to a QHBoxLayout inside a QVBoxLayout
        hboxLayout = QHBoxLayout()
        hboxLayout.addWidget(tableWidget)
        vboxLayout = QVBoxLayout()
        vboxLayout.addLayout(hboxLayout)
        self.ui.tabWidget.widget(self.ui.tabWidget.count() - 1).setLayout(vboxLayout)

        # Change tabs name
        self.ui.tabWidget.setTabText(self.ui.tabWidget.count() - 1, sheet_name)

        # set value of table from data
        headers = list(data.keys())
        tableWidget.setHorizontalHeaderLabels(headers)

        for i, row in enumerate(data):
            for j, header in enumerate(headers):
                row_data = data[header][i]  # Get the dictionary for the current row
                cell_data = list(row_data.values())[0]
                tableWidget.setItem(i, j, QTableWidgetItem(str(cell_data)))
        
        # resize columns to fit the contents
        tableWidget.resizeColumnsToContents()
    
    def handleCellChanged(self, row, column):
        tableWidget = self.sender()
        cell_id = tableWidget.item(row, 0).text()
        
        # send a post request to the server to update the cell
        sheet_name = self.ui.tabWidget.tabText(self.ui.tabWidget.currentIndex())
        headers = network.get_headers(DATA, sheet_name)
        if "id" in headers:
            headers.remove("id")
        # replace ID with row
        # Check if "ID" is in headers
        if "ID" in headers:
            index = headers.index("ID")
            headers[index] = "row"
        
        data = {
            "sheetname": sheet_name,
            "data": {headers[j]: tableWidget.item(row, j).text() for j in range(tableWidget.columnCount())}
        }
        
        response = network.post_data(
            f"{URL}/update/", data
        )
        
        if response.status_code == 200:
            self.loadJsonData(URL, last_page=True)
        else:
            self.showAlarm("Error", "Failed to update the cell!")
        
    
    def loadReport(self, url, name):
        # Fetch the data
        # global DATA # REMOVED BECAUSE OPERATIONS WAS OVERWRITING GLOBAL DATA
        url = f"{url}/operations/{name}"
        try:
            data = network.get_data(url)
            DATA = data
        except Exception as e:
            print(f"Error: {e}")
            self.showAlarm("Network error", "Failed to fetch data from the server!")
            return
        
        if data is None:
            return

        # Display the data in a new table popup
        dialog = QDialog(self)
        dialog.setWindowTitle(name)
        dialog.resize(800, 600)
        layout = QVBoxLayout(dialog)
        table = TableViewer(dialog)
        try:
            if name == "legal-aid":
                table.set_data(data, name="legal-aid")
            elif name == "bail-refused":
                table.set_data(data, name="bail-refused")
            elif name == "stage":
                if table.set_data(data, name="stage") == False:
                    self.showAlarm("Error", "No data to display!")
                    return
            else:
                if table.set_data(data) == False:
                    self.showAlarm("Error", "No data to display!")
                    return
        except IndexError:
            self.showAlarm("Error", "No data to display!")
            return
        layout.addWidget(table)
        dialog.exec()
        
        
    
    def showAlarm(self, header, mes):
        QMessageBox.warning(self, header, mes)

    def showSuccess(self, header, mes):
        QMessageBox.information(self, header, mes)

    def openFile(self):
        self.loadJsonData(URL)

        # connect tables to saveExcelData function
        # self.tableWidgetCellChange(is_connect=True)

    def removeEmptyColumns(self, sheet):
        columns_to_remove = []
        for i, column in enumerate(sheet.iter_cols(values_only=True), start=1):
            if all(cell is None for cell in column):
                columns_to_remove.append(i)

        for i in reversed(columns_to_remove):
            sheet.delete_cols(i)

    def saveExcelData(self):
        # Disconnect the cellChanged signal
        self.tableWidgetCellChange(is_connect=False)

        # Save data into excel file
        for sheet_index in range(self.sheet_number):
            tableWidget = self.ui.tabWidget.widget(sheet_index).findChild(QTableWidget)
            # Write the data from the tableWidget back to the sheet
            for i in range(tableWidget.rowCount()):
                for j in range(tableWidget.columnCount()):
                    if tableWidget.item(i, j) is not None:
                        # seprate logic of first sheet for Writing
                        row_offset = 17 if sheet_index == 0 else 1
                        self.wb[self.wb.sheetnames[sheet_index]].cell(
                            row=i + 1 + row_offset,
                            column=j + 1,
                            value=tableWidget.item(i, j).text(),
                        )
        self.wb.save(self.excel_file)

        # Reconnect the cellChanged signal
        self.tableWidgetCellChange(is_connect=True)

    def askForNewEntry(self, data) -> bool:
        # open a popup window for new entry
        dialog = NewEntryDialog(data, self)
        if dialog.exec():
            selected_sheet = dialog.comboBox.currentText()
            new_entry_data = {}
            for i, lineEdit in enumerate(dialog.lineEdits):
                new_entry_data[
                    dialog.lineEditsLayout.itemAt(i).widget().placeholderText()
                ] = lineEdit.text()
            for entry in new_entry_data:
                # fix showing None value cell
                if new_entry_data[entry].strip() == "":
                    new_entry_data[entry] = None


            new_entry = {
                "sheetname": selected_sheet,
                "data": new_entry_data
            }

            print(new_entry)
            # Send a POST request to the server with the new entry data
            response = network.post_data(
                f"{URL}/create/", new_entry
            )

            if response.status_code == 200:
                self.showSuccess("Success", "New entry added successfully!")
                # refresh data
                self.loadJsonData(URL, last_page=True)
                return True
            else:
                self.showAlarm("Error", "Failed to add new entry!")
                return False

        return False

    def genDocsBtn(self, data):
        if gen_docs(data):
            QMessageBox.information(
                self,
                "Success",
                "Word documents successfully generated! Files are placed in the Docs folder.",
            )
        else:
            self.showAlarm("Error", "Word documents generation failed!")

    def showNewEntryDialog(self):
        global DATA
        data = DATA
        if not self.askForNewEntry(data):
            return
        print("New entry added")

    def showOperationDialog(self):
        dialog = OperationsDialog(self, URL)
        dialog.exec()
    
    def showDeleteEntryDialog(self):
        dialog = DeleteEntryDialog(self, data=DATA)
        dialog.exec()

    def tableWidgetCellChange(self, is_connect: bool):
        """
        connect and disconnect trigger for change cell
        """
        try:
            for i in range(self.sheet_number):
                tableWidget = self.ui.tabWidget.widget(i).findChild(QTableWidget)
                if is_connect:
                    tableWidget.cellChanged.connect(self.saveExcelData)
                else:
                    tableWidget.cellChanged.disconnect(self.saveExcelData)
        except Exception as e:
            print(f"have error when is_connect is {is_connect} in :{e} ")

    def exportToJson(self):
        json_data = {}
        for i in range(self.ui.tabWidget.count()):
            tableWidget = self.ui.tabWidget.widget(i).findChild(QTableWidget)
            sheet_name = self.ui.tabWidget.tabText(i)
            sheet_data = {}
            for j in range(tableWidget.columnCount()):
                header = tableWidget.horizontalHeaderItem(j).text()
                column_data = []
                for k in range(tableWidget.rowCount()):
                    cell_data = tableWidget.item(k, j).text()
                    column_data.append({f"row {k + 1}": cell_data})
                sheet_data[header] = column_data
            json_data[sheet_name] = {"data": sheet_data}
        return json_data

    def closeApplication(self):
        self.close()

class NewEntryDialog(QDialog):
    def __init__(self, data, parent=None):
        super().__init__(parent)
        self.setWindowTitle("New Entry")

        self.setStyleSheet(
            """
            QDialog {
                background-color: #f0f0f0;
            }

            QLabel {
                font-size: 14px;
            }

            QLineEdit, QComboBox, QDateEdit {
                background-color: #fff;
                border: 1px solid #999;
                padding: 5px;
            }

            QPushButton {
                background-color: #007BFF;
                color: #fff;
                border: none;
                padding: 5px 10px;
                margin: 10px;
            }

            QPushButton:hover {
                background-color: #0056b3;
            }
            
            QComboBox:!editable, QComboBox::drop-down:editable {
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                            stop: 0 #E1E1E1, stop: 0.4 #DDDDDD,
                                            stop: 0.5 #D8D8D8, stop: 1.0 #D3D3D3);
            }
        """
        )
        
        main_data = data
        headers = list(data.get("headers"))
        data = data.get("data")
    
        # Set the size of the dialog to be 3/4 of the size of the parent
        if parent is not None:
            self.resize(parent.size() * 0.5)

        self.layout = QVBoxLayout(self)

        self.comboBox = QComboBox(self)
        self.comboBox.addItems(headers)
        self.comboBox.setCurrentText(
            parent.ui.tabWidget.tabText(parent.ui.tabWidget.currentIndex())
        )

        self.layout.addWidget(self.comboBox)

        self.lineEditsLayout = QGridLayout()
        self.layout.addLayout(self.lineEditsLayout)

        self.lineEdits = []
        self.updateLineEdits(main_data.get("headers")[self.comboBox.currentText()])
        

        self.comboBox.currentIndexChanged.connect(
            lambda: self.updateLineEdits(main_data.get("headers")[self.comboBox.currentText()])
        )

        # Add a stretchable space
        self.layout.addStretch(1)

        self.button = QPushButton("Submit", self)
        # change size of button
        self.button.setMinimumSize(100, 60)
        self.button.clicked.connect(self.accept)
        self.layout.addWidget(self.button)

    def updateLineEdits(self, columns):
        # Remove existing QLineEdit widgets
        for lineEdit in self.lineEdits:
            self.lineEditsLayout.removeWidget(lineEdit)
            lineEdit.deleteLater()
        self.lineEdits.clear()

        # Filter out "id" and "row" columns
        columns = [column for column in columns if column not in ["id", "row", "ID"]]

        # Create a new QGridLayout
        self.lineEditsLayout = QGridLayout()

        # Add new QLineEdit widgets
        for i, column in enumerate(columns):
            lineEdit = QLineEdit(self)
            lineEdit.setPlaceholderText(column)
            self.lineEditsLayout.addWidget(lineEdit, i // 4, i % 4)
            self.lineEdits.append(lineEdit)

        # Add the new QGridLayout to the dialog's layout
        self.layout.insertLayout(1, self.lineEditsLayout)
class DeleteEntryDialog(QDialog):
    def __init__(self, parent=None, data=None):
        super().__init__(parent)
        self.setWindowTitle("Delete Entry")

        self.setStyleSheet(
            """
            QDialog {
                background-color: #f0f0f0;
            }

            QLabel {
                font-size: 14px;
            }

            QLineEdit, QComboBox, QDateEdit {
                background-color: #fff;
                border: 1px solid #999;
                padding: 5px;
            }

            QPushButton {
                background-color: #007BFF;
                color: #fff;
                border: none;
                padding: 5px 10px;
                margin: 10px;
            }

            QPushButton:hover {
                background-color: #0056b3;
            }
            
            QComboBox:!editable, QComboBox::drop-down:editable {
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                            stop: 0 #E1E1E1, stop: 0.4 #DDDDDD,
                                            stop: 0.5 #D8D8D8, stop: 1.0 #D3D3D3);
            }
        """
        )
        
        main_data = data
        headers = list(data.get("headers"))
        data = data.get("data")
    
        # Set the size of the dialog to be 3/4 of the size of the parent
        if parent is not None:
            self.resize(parent.size() * 0.5)

        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(10, 5, 10, 10)  # Reduce the top margin

        # Add QLabel for sheet selection
        self.sheetLabel = QLabel("Select Sheet:", self)
        self.layout.addWidget(self.sheetLabel)

         # Add QComboBox for sheet selection
        self.sheetComboBox = QComboBox(self)
        self.sheetComboBox.addItems(headers)
        self.sheetComboBox.setCurrentText(
            parent.ui.tabWidget.tabText(parent.ui.tabWidget.currentIndex())
        )
        self.layout.addWidget(self.sheetComboBox)

        # Add QLabel for row number
        self.rowLabel = QLabel("ID Number:", self)
        self.layout.addWidget(self.rowLabel)

        # Add QLineEdit for row number
        self.rowLineEdit = QLineEdit(self)
        self.rowLineEdit.setPlaceholderText("Enter row number")
        self.layout.addWidget(self.rowLineEdit)
        
        # Add a stretchable space
        self.layout.addStretch(1)
        
        # Add QPushButton for submitting the form
        self.button = QPushButton("Delete", self)
        self.button.setMinimumSize(100, 60)
        self.button.clicked.connect(lambda: self.deleteEntry(parent, data))
        self.layout.addWidget(self.button)
        
        # auto resize the dialog
        self.adjustSize()


    def deleteEntry(self, parent, data):
        # Get the selected sheet name
        sheet_name = self.sheetComboBox.currentText()
        
        # Get the row number
        row_number = self.rowLineEdit.text()
        
        # Get the data of the selected sheet
        sheet_data = data.get(sheet_name)

        # Extract all IDs from the sheet data
        ids = [entry['ID'] for entry in sheet_data]

        # Check if the row number is valid
        if not row_number.isdigit():
            parent.showAlarm("Error", "Please enter a valid ID number!")
            return

        # Convert the row number to an integer
        row_number = int(row_number)

        # Check if the ID exists in the list of IDs
        if row_number not in ids:
            parent.showAlarm("Error", "ID number does not exist!")
            return

        request_data = {
            "sheetname": sheet_name,
            "row": row_number
        }
        print(request_data)

        # Send a DELETE request to the server to delete the entry
        response = network.post_data(
            f"{URL}/delete/", request_data
        )
        
        if response.status_code == 200:
            parent.showSuccess("Success", "Entry deleted successfully!")
            # Refresh the data
            parent.loadJsonData(URL, last_page=True)
        else:
            parent.showAlarm("Error", "Failed to delete the entry!")
            print(response.text)
        
        # Close the dialog
        self.close()

class OperationsDialog(QDialog):
    def __init__(self, parent, url):
        super().__init__()

        self.setStyleSheet(
            """
            QPushButton {
                background-color: #007cff; /* Green */
                border: none;
                color: white;
                padding: 15px 32px;
                text-align: center;
                text-decoration: none;
                font-size: 16px;
                margin: 4px 2px;
            }
            
            QPushButton:hover {
                background-color: #3094fd;
            }
        """
        )

        self.parent = parent

        self.setWindowTitle("Operations")

        self.layout = QGridLayout(self)

        self.monthly_cases_report_button = QPushButton("Cases this Month")
        self.monthly_cases_report_button.setMinimumSize(100, 40)
        self.layout.addWidget(self.monthly_cases_report_button, 0, 0)
        self.monthly_cases_report_button.clicked.connect(
            lambda: (parent.loadReport(url, "monthly"), self.accept())
        )

        self.weekly_cases_report_button = QPushButton("Cases this Week")
        self.weekly_cases_report_button.setMinimumSize(100, 40)
        self.layout.addWidget(self.weekly_cases_report_button, 0, 1)
        self.weekly_cases_report_button.clicked.connect(
            lambda: (parent.loadReport(url, "weekly"), self.accept())
        )

        self.legal_aid_report_button = QPushButton("Legal Aid")
        self.legal_aid_report_button.setMinimumSize(100, 40)
        self.layout.addWidget(self.legal_aid_report_button, 1, 0)
        self.legal_aid_report_button.clicked.connect(
            lambda: (parent.loadReport(url, "legal-aid"), self.accept())
        )

        self.bail_refused_report_button = QPushButton("Bail Refused")
        self.bail_refused_report_button.setMinimumSize(100, 40)
        self.layout.addWidget(self.bail_refused_report_button, 1, 1)
        self.bail_refused_report_button.clicked.connect(
            lambda: (parent.loadReport(url, "bail-refused"), self.accept())
        )

        self.empty_counsel_report_button = QPushButton("Empty Counsel")
        self.empty_counsel_report_button.setMinimumSize(100, 40)
        self.layout.addWidget(self.empty_counsel_report_button, 2, 0)
        self.empty_counsel_report_button.clicked.connect(
            lambda: (parent.loadReport(url, "empty-counsel"), self.accept())
        )

        self.non_zero_balance_report_button = QPushButton("Non Zero Balance")
        self.non_zero_balance_report_button.setMinimumSize(100, 40)
        self.layout.addWidget(self.non_zero_balance_report_button, 2, 1)
        self.non_zero_balance_report_button.clicked.connect(
            lambda: (parent.loadReport(url, "non-zero"), self.accept())
        )

        self.stage_reports_button = QPushButton("Stage Reports")
        self.stage_reports_button.setMinimumSize(100, 40)
        self.layout.addWidget(self.stage_reports_button, 3, 0)
        self.stage_reports_button.clicked.connect(
            lambda: (parent.loadReport(url, "stage"), self.accept())
        )

class FetchDataThread(QThread):
    dataReady = Signal(object)

    def __init__(self, url):
        super().__init__()
        self.url = url

    def run(self):
        try:
            data = network.get_data(self.url)
            
            global DATA
            DATA = data
            
            self.dataReady.emit(data)
        except Exception as e:
            print(f"Error: {e}")
            self.dataReady.emit("Error")


class LoadingDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setModal(True)
        self.setWindowTitle("Loading...")
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowCloseButtonHint)
        # self.setWindowFlags(Qt.Dialog | Qt.CustomizeWindowHint | Qt.Tool)
        self.setStyleSheet(
            """
            QDialog {
                background-color: #fff;
                color: #000;
                border: 1px solid #999;
            }
        """
        )

        layout = QVBoxLayout()
        label = QLabel("Please wait...")
        label.setAlignment(Qt.AlignCenter)
        layout.addWidget(label)

        progress = QProgressBar(self)
        progress.setRange(
            0, 0
        )  # Set range to 0,0 to create an indeterminate progress bar
        layout.addWidget(progress)

        self.setLayout(layout)

class TableViewer(QTableWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setSortingEnabled(True)
        self.sortItems(0)

    def set_data(self, data, name=None):
        # Get the headers from the first row (the keys of the dictionary)
        if name == "legal-aid":
            headers = list(data.keys())
            self.setColumnCount(2)
            self.setHorizontalHeaderLabels(["Legal Aid Category", "Number of Clients"])
        elif name == "bail-refused":
            table_headers = ["Office", "Matter_Type", "Email", "File_No", "Fee_Earner", "Clients_Surname",
                            "Clients_Forenames", "Address", "City", "Postcode", "Mobile_Number", "Date_of_Birth",
                            "HMP", "Prison_Number", "National_Insurance_Number", "Legal_Aid", "Court"]
            self.setColumnCount(len(table_headers))
            self.setHorizontalHeaderLabels(table_headers)
        elif name == "stage":
            # check to see if all the keys under data are empty if yes then return false
            if all([data[key] == [] for key in data]):
                return False
            headers = list(data[0].keys())
            self.setColumnCount(len(headers))
            # Set the headers
            self.setHorizontalHeaderLabels(headers)
        else:
            headers = list(data[0].keys())
            self.setColumnCount(len(headers))
            # Set the headers
            self.setHorizontalHeaderLabels(headers)
        
        if name == "bail-refused":
            row_count = 0
            for value in data.values():
                row_count += len(value)
            self.setRowCount(row_count)
        else:
            self.setRowCount(len(data))
        
        
        # Set the data
        if name == "legal-aid":
            for i, header in enumerate(headers):
                cell_data = data[header]
                if header == "__null__":
                    header = ""
                    cell_data = ""
                if isinstance(cell_data, float):
                    cell_data = int(cell_data)
                if cell_data == "__null__":
                    cell_data = ""
                self.setItem(i, 0, QTableWidgetItem(header))
                self.setItem(i, 1, QTableWidgetItem(str(cell_data)))
        elif name == "bail-refused":
            row_count = 0
            for value in data.values():
                for dic in value:
                    for j, header in enumerate(table_headers):
                        cell_data = dic[header]
                        if isinstance(cell_data, float):
                            cell_data = int(cell_data)
                        if cell_data == "__null__":
                            cell_data = ""
                        self.setItem(row_count, j, QTableWidgetItem(str(cell_data)))
                    row_count += 1
        else:
            for i, row in enumerate(data):
                for j, header in enumerate(headers):                    
                    cell_data = row[header]
                    if isinstance(cell_data, float):
                        cell_data = int(cell_data)
                    if cell_data == "__null__":
                        cell_data = ""
                    self.setItem(i, j, QTableWidgetItem(str(cell_data)))
        
        # Resize the columns to fit the contents
        self.resize_columns()
        
    def resize_columns(self):
        self.resizeColumnsToContents()

def run():
    app = QApplication(sys.argv)
    widget = MainWindow()
    widget.openFile()
    widget.showMaximized()
    sys.exit(app.exec())


if __name__ == "__main__":
    run()
