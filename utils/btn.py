import openpyxl
from datetime import datetime, timedelta
from pprint import pprint


def generate_bail_refused_report(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path)

    magistrates_sheet = workbook["Magistrates_Merge"]
    crown_court_sheet = workbook["Crown Court Merge"]
    # report_sheet = workbook.get_sheet_by_name(
    #     "Clients in Prison Report"
    # )  # Get existing or create new

    # if not report_sheet:
    report_sheet = workbook.create_sheet("Clients in Prison Report")

    bail_col = get_column_index(magistrates_sheet, "Bail")
    prison_number_col = get_column_index(magistrates_sheet, "Prison Number")
    file_no_col = get_column_index(magistrates_sheet, "File No.")

    copy_headers(magistrates_sheet, report_sheet)

    report_row = 2  # Start from row 2 to leave space for headers
    unique_file_numbers = set()

    extract_data_for_bail_refused(
        magistrates_sheet,
        report_sheet,
        report_row,
        bail_col,
        prison_number_col,
        file_no_col,
        unique_file_numbers,
    )
    extract_data_for_bail_refused(
        crown_court_sheet,
        report_sheet,
        report_row,
        bail_col,
        prison_number_col,
        file_no_col,
        unique_file_numbers,
    )

    workbook.save(workbook_path)
    print("Bail refused report generated successfully!")


def extract_data_for_bail_refused(
    data_sheet,
    report_sheet,
    report_row,
    bail_col,
    prison_number_col,
    file_no_col,
    unique_file_numbers,
):
    criteria = ["Bail Refused", "JC Bail Refused"]

    for row in data_sheet.iter_rows(min_row=2):  # Skip header row
        bail_value = row[bail_col - 1].value
        prison_number = row[prison_number_col - 1].value
        file_no = row[file_no_col - 1].value

        if (
            bail_value in criteria
            and prison_number is not None
            and file_no not in unique_file_numbers
        ):
            unique_file_numbers.add(file_no)
            for i in range(1, 18):  # Copy 17 columns
                report_sheet.cell(row=report_row, column=i).value = row[i - 1].value
            report_row += 1


def get_column_index(sheet, header):
    header_row = next(sheet.iter_rows(min_row=1, max_row=1))  # Get first row
    for cell in header_row:
        if cell.value == header:
            return cell.column
    return 0


def copy_headers(source_sheet, target_sheet):
    for i in range(1, 18):  # Copy 17 headers
        target_sheet.cell(row=1, column=i).value = source_sheet.cell(
            row=1, column=i
        ).value


####################
def create_upcoming_month_sheet(workbook_path):
    wb = openpyxl.load_workbook(workbook_path)

    # Access the "Opening File" worksheet
    ws_source = wb["Opening File"]

    # Check if "Upcoming Cases this Month" sheet exists, if not create it
    if "Upcoming Cases this Month" not in wb.sheetnames:
        wb.create_sheet("Upcoming Cases this Month")
    ws_target = wb["Upcoming Cases this Month"]

    # Assuming the first row is the header and we're filtering based on a condition in column 1
    # Copy header row to target sheet
    for col in range(1, ws_source.max_column + 1):
        ws_target.cell(row=1, column=col).value = ws_source.cell(
            row=1, column=col
        ).value

    target_row = 1  # Start writing to the second row of the target sheet
    # data of first sheet will start from 17 row
    for index, row in enumerate(ws_source.iter_rows(min_row=17, values_only=True)):
        # Apply filtering condition - for example, checking a date or a specific text
        # This is where you'd customize based on your actual filter condition
        if index == 0:
            print("yes it is ")
            for col, value in enumerate(row, start=1):
                ws_target.cell(row=target_row, column=col).value = value
            target_row += 1
        elif date_open_filter(row) and row:
            for col, value in enumerate(row, start=1):
                ws_target.cell(row=target_row, column=col).value = value
            target_row += 1

    # Save the workbook with the new data
    wb.save("upcommit_month.xlsx")


def create_upcoming_week_sheet(workbook_path):
    wb = openpyxl.load_workbook(workbook_path)

    # Access the "Opening File" worksheet
    ws_source = wb["Opening File"]

    # Check if "Upcoming Cases this Month" sheet exists, if not create it
    if "Upcoming Cases this Month" not in wb.sheetnames:
        wb.create_sheet("Upcoming Cases this Month")
    ws_target = wb["Upcoming Cases this Month"]

    # Assuming the first row is the header and we're filtering based on a condition in column 1
    # Copy header row to target sheet
    for col in range(1, ws_source.max_column + 1):
        ws_target.cell(row=1, column=col).value = ws_source.cell(
            row=1, column=col
        ).value

    target_row = 1  # Start writing to the second row of the target sheet
    # data of first sheet will start from 17 row
    for index, row in enumerate(ws_source.iter_rows(min_row=17, values_only=True)):
        # Apply filtering condition - for example, checking a date or a specific text
        # This is where you'd customize based on your actual filter condition
        if index == 0:
            print("yes it is ")
            for col, value in enumerate(row, start=1):
                ws_target.cell(row=target_row, column=col).value = value
            target_row += 1
        elif date_open_filter(row) and row:
            for col, value in enumerate(row, start=1):
                ws_target.cell(row=target_row, column=col).value = value
            target_row += 1

    # Save the workbook with the new data
    wb.save("upcommit_month.xlsx")


def create_legal_aid_report_sheet(workbook_path):
    def calc_report_vars(last_column_values):
        vals = {
            "Submitted": 0,
            "Refused": 0,
            "Date Stamped": 0,
            "Approved": 0,
            "Appealed": 0,
        }
        for i in last_column_values:
            if i and i in vals:
                vals[i] += 1
        return vals

    # Load the Excel file
    wb = openpyxl.load_workbook(workbook_path)

    # Select the specific sheet
    sheet = wb["Opening File"]  # Replace 'Sheet1' with the name of your sheet

    # Find the last column with data
    last_column = sheet.max_column

    # Create an empty list to store values from the last column
    last_column_values = [
        sheet.cell(row=row, column=last_column).value
        for row in range(18, sheet.max_row + 1)
    ]
    print(calc_report_vars(last_column_values))

    # Close the Excel file
    wb.close()


def is_current_week(date_str):
    # Assuming the date is in 'YYYY-MM-DD' format; adjust the format as necessary
    date_opened = datetime.strptime(date_str, "%Y-%m-%d")

    today = datetime.today()
    start_week = today - timedelta(days=today.isoweekday() - 1)  # Monday
    end_week = start_week + timedelta(days=6)  # Sunday

    return start_week <= date_opened <= end_week


def date_open_filter(row):
    # Assume the "date opened" is in the first column (index 0) of the row
    date_opened_str = row[7]
    print(type(date_opened_str))
    print(date_opened_str)

    # Assuming the date is in 'YYYY-MM-DD' format; adjust the format as necessary
    # date_opened = datetime.strptime(date_opened_str, "%d.%m.%Y")
    date_opened = date_opened_str

    # Get the first and last day of the current month
    today = datetime.today()
    first_day_of_month = datetime(today.year, today.month, 1)
    last_day_of_month = datetime(today.year, today.month + 1, 1) - timedelta(days=1)

    # Check if the date_opened falls within the current month
    return first_day_of_month <= date_opened <= last_day_of_month


def clear_worksheet(workbook_path, sheet_name):
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb[sheet_name]

    # Check if the sheet has any data to clear
    if ws.max_row > 0:
        # ws.delete_rows() can be used to delete rows in the given range
        ws.delete_rows(1, ws.max_row)

    # Save the workbook after clearing the sheet
    wb.save(workbook_path)


def get_sheet_names(workbook):
    """
    Returns a list of sheet names from the provided openpyxl workbook object.

    Args:
        workbook (openpyxl.Workbook): The openpyxl workbook object.

    Returns:
        list: A list containing the names of all sheets in the workbook.
    """
    return workbook.sheetnames


excel_file = "../Law Clients Excel Sheet Shared_MainV3.xlsx"

generate_bail_refused_report(excel_file)
