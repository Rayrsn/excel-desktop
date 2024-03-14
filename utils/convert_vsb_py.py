import openpyxl
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from pprint import pprint

# from pprint import pprint


def get_column_val(
    column_num,
    workbook,
    data_worksheet_name="Opening File",
    start_row=17,
):
    """
    get value of column that are not None and save them in array
    output:
        (
            header,
            [(index, value), ...]
        )

    """

    data_worksheet = workbook[data_worksheet_name]
    values = []
    for index, row in enumerate(
        data_worksheet.iter_rows(min_row=start_row, values_only=True)
    ):
        if index == 0:
            header = row[column_num]
            continue
        if row[column_num] is None:
            continue
        values.append((index, row[column_num]))
    return (header, values)


def first_sh_rows_with_numbers(filepath):
    """Reads rows from the first sheet and saves them in a list with row numbers."""

    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook["Opening File"]

    data_with_row_numbers = []
    # 17 row is header
    for row_num, row in enumerate(
        sheet.iter_rows(min_row=17),
        start=17 + 1,  # add 1 because of row in excel file start with 1
    ):
        # filter empty row with first column
        if row[0].value == None:
            continue
        # Skip the header row (row 1)
        data_list = [cell.value for cell in row]
        data_with_row_numbers.append((row_num, data_list))

    return data_with_row_numbers


def create_sheet(workbook, sheet_name):
    """create sheet and return it"""
    try:
        report_worksheet = workbook[sheet_name]

        # clear sheet
        for row in report_worksheet.iter_rows():
            for cell in row:
                cell.value = None
    except:
        report_worksheet = workbook.create_sheet(sheet_name)
    return report_worksheet


def write_header(
    workbook,
    target_sh_name,
    source_sh="Opening File",
    start_header_row=17,
    header_list=None,
) -> bool:
    """
    Write header into target sheeet
    it work with get source sheet or header list
    """

    report_worksheet = workbook[target_sh_name]

    # Get references to worksheets
    if source_sh:
        data_worksheet = workbook[source_sh]
        # Filter data for the current month
        header_row = data_worksheet[start_header_row]
        for index, header in enumerate(header_row):
            report_worksheet.cell(
                row=0 + 1, column=index + 1, value=header_row[index].value
            )
    elif header_list and not source_sh:
        for index, header_item in enumerate(header_list):
            report_worksheet.cell(row=0 + 1, column=index + 1, value=header_item)
    else:
        return False

    return True


def write_rows(filepath, target_sh, rows, start_row=2, is_first_sh_filter=False):
    """Write rows rows list"""
    # NOTE : this fucntion write for generate_monthly_cases_report and it filter

    if is_first_sh_filter:
        f_sh_rows = first_sh_rows_with_numbers(filepath)
        for row_num, row in rows:
            for f_sh_row_num, row in f_sh_rows:
                if f_sh_row_num == row_num:
                    for col, cell in enumerate(row):
                        target_sh.cell(row=start_row, column=col + 1, value=cell)
                    start_row += 1
    else:
        for row_num, row in enumerate(rows):
            for col, cell in row:
                target_sh.cell(row=start_row, column=col + 1, value=cell)
            start_row += 1


def generate_monthly_cases_report(filepath):
    """Generates a monthly cases report from an Excel workbook."""

    workbook = openpyxl.load_workbook(filepath)

    # Get references to worksheets
    data_worksheet = workbook["Opening File"]
    report_worksheet_name = "Upcoming Cases this Month"

    # create report sheet
    report_worksheet = create_sheet(workbook, report_worksheet_name)

    current_month_start = datetime.now().replace(day=1)
    current_month_end = (current_month_start + relativedelta(months=+1)) - timedelta(
        days=1
    )
    _, date_opened_vals = get_column_val(7, workbook)
    filtered_data = [
        (index + 18, cell)  # add 17 for 17 row was skip
        for index, cell in date_opened_vals
        if current_month_start <= cell <= current_month_end
    ]

    # copy header into report worksheet
    write_header(workbook=workbook, target_sh_name=report_worksheet_name)

    # copy filtered_data into report sheet
    write_rows(filepath, target_sh=report_worksheet, rows=filtered_data)

    workbook.save(filepath)

    print("Monthly cases report generated successfully!")


def generate_weekly_cases_report(filepath):
    """Generates a weekly cases report from an Excel workbook."""

    workbook = openpyxl.load_workbook(filepath)

    # Get references to worksheets
    data_worksheet = workbook["Opening File"]
    report_worksheet_name = "Upcoming Cases this Week"

    # create report sheet
    report_worksheet = create_sheet(workbook, report_worksheet_name)

    # Calculate start and end dates for the current week
    today = datetime.today()
    weekday = today.weekday()  # 0 for Monday, 6 for Sunday
    start_date = today - timedelta(days=weekday)  # Monday of the current week
    end_date = start_date + timedelta(days=6)  # Sunday of the current week

    # Filter data for the current week
    _, date_opened_vals = get_column_val(7, workbook)
    filtered_data = [
        (index + 18, cell)  # add 17 for 17 row was skip
        for index, cell in date_opened_vals
        if start_date <= cell <= end_date
    ]

    # copy header into report worksheet
    write_header(workbook=workbook, target_sh_name=report_worksheet_name)

    # copy filtered_data into report sheet
    write_rows(filepath, target_sh=report_worksheet, rows=filtered_data)

    # Save the workbook
    workbook.save(filepath)

    print("Weekly cases report generated successfully!")


def generate_legal_aid_report(filepath):
    """Generates a legal aid report from an Excel workbook."""

    def calc_report_vars(legal_aid_col_list):
        vals = {
            "Submitted": 0,
            "Refused": 0,
            "Date Stamped": 0,
            "Approved": 0,
            "Appealed": 0,
        }
        for index, cell in legal_aid_col_list:
            if cell and cell in vals:
                vals[cell] += 1
        return vals

    workbook = openpyxl.load_workbook(filepath)

    # Get references to worksheets
    data_worksheet = workbook["Opening File"]
    report_worksheet_name = "Legal Aid Report"

    # create report sheet
    report_worksheet = create_sheet(workbook, report_worksheet_name)

    # Get references to tables
    try:
        header, legal_aid_col = get_column_val(36, workbook)
    except:
        # Check if legal aid column exists
        print("Error: Legal Aid column not found in MasterData table!")
        return

    cell_vals = calc_report_vars(legal_aid_col)

    header = [
        "Legal Aid Category",
        "Submitted",
        "Refused",
        "Date Stamped",
        "Approved",
        "Appealed",
        "Total Number of Clients",
    ]

    client_count = 0
    for _, val in cell_vals.items():
        client_count += val

    result_row = [
        [
            (0, "Number of Clients"),
            (1, cell_vals["Submitted"]),
            (2, cell_vals["Refused"]),
            (3, cell_vals["Date Stamped"]),
            (4, cell_vals["Approved"]),
            (5, cell_vals["Appealed"]),
            (6, client_count),
        ]
    ]

    # Write data to report worksheet
    write_header(
        workbook=workbook,
        target_sh_name=report_worksheet_name,
        source_sh="",
        start_header_row=2,
        header_list=header,
    )

    write_rows(filepath=filepath, target_sh=report_worksheet, rows=result_row)

    # Save the workbook
    workbook.save(filepath)

    print("Legal aid report generated successfully!")


def generate_bail_refused_report(filepath):
    """Generates a bail refused report from an Excel workbook."""

    workbook = openpyxl.load_workbook(filepath)

    # Get references to worksheets
    magistrates_sheet = workbook["Magistrates Merge"]
    crown_court_sheet = workbook["Crown Court Merge"]
    report_sheet_name = "Clients in Prison Report"
    report_sheet = workbook.create_sheet(report_sheet_name)
    # report_sheet = workbook.get_sheet_by_name(report_sheet_name)
    # if not report_sheet:
    # report_sheet = workbook.create_sheet(report_sheet_name)
    # else:
    for row in report_sheet.iter_rows():
        for cell in row:
            cell.value = None  # Clear existing data

    # Define column indices (assuming headers are in the first row)
    bail_col = get_column_index(magistrates_sheet, 1, "Bail")
    prison_number_col = get_column_index(magistrates_sheet, 1, "Prison Number")
    file_no_col = get_column_index(magistrates_sheet, 1, "File No.")

    # Copy headers from Magistrates Merge sheet to report sheet
    for i in range(1, 18):  # Assuming 17 headers (A-Q)
        report_sheet.cell(row=1, column=i).value = magistrates_sheet.cell(
            row=1, column=i
        ).value

    # Initialize variables
    report_row = 2  # Start from row 2 for data
    used_file_numbers = set()  # Use a set to store unique file numbers

    # Extract data from Magistrates Merge sheet
    extract_data_for_bail_refused(
        magistrates_sheet.iter_rows(min_row=2),
        report_sheet,
        report_row,
        bail_col,
        prison_number_col,
        file_no_col,
        used_file_numbers,
    )

    # Extract data from Crown Court Merge sheet
    extract_data_for_bail_refused(
        crown_court_sheet.iter_rows(min_row=2),
        report_sheet,
        report_row,
        bail_col,
        prison_number_col,
        file_no_col,
        used_file_numbers,
    )

    # Create a table from the report data
    report_table = openpyxl.worksheet.table.Table(
        displayName="Clients in Prison Report",
        ref="A1:" + report_sheet.cells(report_sheet.max_row, 1).end(xlUp).row,
    )
    report_table.table_style = "TableStyleLight9"  # Optional table style
    report_sheet.add_table(report_table)

    # Save the workbook
    workbook.save(filepath)

    print("Bail refused report generated successfully!")


def extract_data_for_bail_refused(
    data_rows,
    report_sheet,
    report_row,
    bail_col,
    prison_number_col,
    file_no_col,
    used_file_numbers,
):
    """Extracts data for bail refused cases and writes to report sheet."""

    bail_criteria = ("Bail Refused", "JC Bail Refused")

    for row in data_rows:
        cells = [cell.value for cell in row]
        if (
            cells[bail_col - 1] in bail_criteria
            and cells[prison_number_col - 1]
            and cells[file_no_col - 1] not in used_file_numbers
        ):
            used_file_numbers.add(cells[file_no_col - 1])
            for i in range(len(cells)):
                report_sheet.cell(row=report_row, column=i + 1).value = cells[i]
            report_row += 1


def get_column_index(worksheet, header_row, header_text):
    """Gets the column index of a header in a worksheet."""

    for cell in worksheet.iter_rows(min_row=header_row):
        for col, value in enumerate(cell, 1):
            if value == header_text:
                return col
    return 0  # Return 0 if header not found


# def get_column_index(worksheet, header):
#     """Gets the column index of a header in a worksheet."""

#     for col, value in enumerate(worksheet.iter_rows(min_row=1), 1):
#         if header in value:
#             return col
#     return 0


def generate_empty_counsel_report(filepath):
    """Generates a report of rows with empty counsel names in the Crown Court sheet."""

    workbook = openpyxl.load_workbook(filepath)

    # Get references to worksheets
    crown_court_sheet = workbook["Crown Court Merge"]
    report_sheet_name = "Empty Counsel Report"
    report_sheet = workbook.create_sheet(report_sheet_name)
    # report_sheet = workbook.get_sheet_by_name(report_sheet_name)
    # if not report_sheet:
    #     report_sheet = workbook.create_sheet(report_sheet_name)
    # else:
    for row in report_sheet.iter_rows():
        for cell in row:
            cell.value = None  # Clear existing data

    # Define counsel column index (assuming headers are in the first row)
    counsel_col = get_column_index(crown_court_sheet, 1, "Name of Counsel")

    # Copy headers from Crown Court sheet to report sheet
    for i in range(1, crown_court_sheet.max_column + 1):
        report_sheet.cell(row=1, column=i).value = crown_court_sheet.cell(
            row=1, column=i
        ).value

    # Initialize report row
    report_row = 2

    # Extract data from Crown Court sheet with empty counsel names
    extract_data_for_empty_counsel(
        crown_court_sheet.iter_rows(min_row=2), report_sheet, report_row, counsel_col
    )

    # Create a table from the report data
    report_table = openpyxl.worksheet.table.Table(
        displayName="Empty Counsel Report",
        ref="A1:" + report_sheet.cells(report_sheet.max_row, 1).end(xlUp).row,
    )
    report_table.table_style = "TableStyleLight9"  # Optional table style
    report_sheet.add_table(report_table)

    # Save the workbook
    workbook.save(filepath)

    print("Empty counsel report generated successfully!")


def extract_data_for_empty_counsel(data_rows, report_sheet, report_row, counsel_col):
    """Extracts rows with empty counsel names and writes them to the report sheet."""

    for row in data_rows:
        cells = [cell.value for cell in row]
        if not cells[counsel_col - 1]:
            for i in range(len(cells)):
                report_sheet.cell(row=report_row, column=i + 1).value = cells[i]
            report_row += 1


def generate_non_zero_balance_report(filepath):
    """Generates a report of rows with non-zero balance in the Road Traffic sheet."""

    workbook = openpyxl.load_workbook(filepath)

    # Get references to worksheets
    road_traffic_sheet = workbook["Road Traffic"]
    report_sheet_name = "Non-Zero Balance Report"
    report_sheet = workbook.create_sheet(report_sheet_name)
    # report_sheet = workbook.get_sheet_by_name(report_sheet_name)
    # if not report_sheet:
    #     report_sheet = workbook.create_sheet(report_sheet_name)
    # else:
    for row in report_sheet.iter_rows():
        for cell in row:
            cell.value = None  # Clear existing data

    # Define balance column index (assuming headers are in the first row)
    balance_col = get_column_index(road_traffic_sheet, 1, "Balance")

    # Copy headers from Road Traffic sheet to report sheet
    for i in range(1, road_traffic_sheet.max_column + 1):
        report_sheet.cell(row=1, column=i).value = road_traffic_sheet.cell(
            row=1, column=i
        ).value

    # Initialize report row
    report_row = 2

    # Extract data from Road Traffic sheet with non-zero balance
    extract_data_for_non_zero_balance(
        road_traffic_sheet.iter_rows(min_row=2), report_sheet, report_row, balance_col
    )

    # Create a table from the report data
    report_table = openpyxl.worksheet.table.Table(
        displayName="Non-Zero Balance Report",
        ref="A1:" + report_sheet.cells(report_sheet.max_row, 1).end(xlUp).row,
    )
    report_table.table_style = "TableStyleLight9"  # Optional table style
    report_sheet.add_table(report_table)

    # Save the workbook
    workbook.save(filepath)

    print("Non-zero balance report generated successfully!")


def extract_data_for_non_zero_balance(data_rows, report_sheet, report_row, balance_col):
    """Extracts rows with non-zero balance and writes them to the report sheet."""

    for row in data_rows:
        cells = [cell.value for cell in row]
        if cells[balance_col - 1] is not None and cells[balance_col - 1] != 0:
            for i in range(len(cells)):
                report_sheet.cell(row=report_row, column=i + 1).value = cells[i]
            report_row += 1


def generate_stage_reports(filepath):
    """Generates stage reports for the current month from the Crown Court Merge sheet."""

    workbook = openpyxl.load_workbook(filepath)
    crown_court_sheet = workbook["Crown Court Merge"]
    current_month = datetime.date.today().month

    # Generate reports for each stage
    stages = ["Stage 1", "Stage 2", "Stage 3", "Stage 4"]
    for stage in stages:
        generate_stage_report(workbook, crown_court_sheet, stage, current_month)

    workbook.save(filepath)
    print("Stage reports generated successfully!")


def generate_stage_report(workbook, ws, stage, current_month):
    """Generates a report for a specific stage and month."""

    report_sheet_name = stage + " Report"
    report_sheet = workbook.get_sheet_by_name(report_sheet_name)
    if not report_sheet:
        report_sheet = workbook.create_sheet(report_sheet_name)
    else:
        for row in report_sheet.iter_rows():
            for cell in row:
                cell.value = None  # Clear existing data

    stage_col = get_column_index(ws, stage)

    # Copy headers
    for i in range(1, ws.max_column + 1):
        report_sheet.cell(row=1, column=i).value = ws.cell(row=1, column=i).value

    report_row = 2

    # Filter and extract data
    for cell in ws.iter_cols(min_row=2, min_col=stage_col, max_col=stage_col):
        for row_cell in cell:
            if (
                isinstance(row_cell.value, datetime.date)
                and row_cell.value.month == current_month
            ):
                extract_data_for_stage_report(
                    row_cell.row, ws, report_sheet, report_row
                )

    # Create a table
    report_table = openpyxl.worksheet.table.Table(
        displayName=stage,
        ref="A1:" + report_sheet.cells(report_sheet.max_row, 1).end("up").row,
    )
    report_table.table_style = "TableStyleLight9"  # Optional table style
    report_sheet.add_table(report_table)

    if report_row == 2:
        print(f"No data found for {stage}")


def extract_data_for_stage_report(row_num, ws, report_sheet, report_row):
    """Copies a row of data to the report sheet."""

    for i in range(1, ws.max_column + 1):
        report_sheet.cell(row=report_row, column=i).value = ws.cell(
            row=row_num, column=i
        ).value
    report_row += 1


if __name__ == "__main__":
    filepath = "../Law Clients_test_month_sheet.sh.xlsx"

    # generate_non_zero_balance_report(filepath)
    # generate_empty_counsel_report(filepath)
    # generate_bail_refused_report(filepath)
    generate_legal_aid_report(filepath)

    # generate_weekly_cases_report(filepath)
    # generate_monthly_cases_report(filepath)
    # generate_stage_reports(filepath)
