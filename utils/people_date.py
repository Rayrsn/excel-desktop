from openpyxl import load_workbook

headers_list = [
    "Sr No.",
    "Office",
    "Matter Type",
    "Email",
    "Previous Number",
    "CRIME",
    "File No.",
    "Date Opened",
    "Fee Earner",
    "Client's Surname",
    "Client's Forename(s)",
    "Client's Title",
    "Marital Status",
    "Letters to Home Address",
    "Address",
    "City",
    "Postcode",
    "Postal Address (if Different)",
    "Postal Address Postcode",
    "Home Telephone",
    "Work Telephone",
    "Mobile Number",
    "Occupation",
    "Date of Birth",
    "Ethnicity",
    "HMP",
    "Prison Number",
    "National Insurance Number",
    "3rd Party",
    "Initial",
    "Conflict",
    "Date",
    "Costs Information",
    "Cost Estimate",
    "Charge Basis",
    "Court",
    "Legal Aid",
]


def last_row(wb, sheet):
    """
    check first column if data doesn't exist it return last row number
    """
    # NOTE: row number 17 is header
    last_row_with_data = 17
    last_row = sheet.max_row
    for row in range(17, last_row):
        selected_row = list(sheet[row])
        # check first header data is Exist or no
        if selected_row[0].value == None:
            last_row_with_data = row
            break
    return last_row_with_data

def people_sheet_data(wb, sheet, row):
    """
    just return people data from row
    """
    # Read the first row
    first_row = [cell.value for cell in sheet[row]]
    people_data = dict(zip(headers_list, first_row))
    return people_data


def get_all_people_data(client_data):
  """
  Extracts client data from a list of dictionaries.

  Args:
      client_data (list): A list of dictionaries containing client data.

  Returns:
      list: A list of dictionaries containing client data formatted for word generation.
  """
  all_data = []
  for client in client_data:
    data = {
      "email": client["Email"],
      "file_no": client["File_No"],
      "date_opened": client["Date_Opened"],
      # ... Extract all other relevant data fields ...
      "matter_type": client["Matter_Type"],
      "m_3rd_party": client["_3rd_Party"],
      "initial": client["Initial"],
    }
    all_data.append(data)
  return all_data



if __name__ == "__main__":
    from pprint import pprint
    excel_file = "Law Clients.xlsm"
    # Load the workbook and select the first sheet
    wb = load_workbook(excel_file)
    # This selects the first sheet
    sheet = wb.worksheets[0]

    print(f"last row is {last_row(wb, sheet)}\n")

    print(f"header number is {len(headers_list)}\n")

    print(f"all people sheet data is: \n")

    pprint(get_all_people_data(wb, sheet))
