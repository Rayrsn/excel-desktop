import pandas as pd
import openpyxl
import zipfile

######################
# +---------+
# |  Debug  |
# +---------+

from colorama import Fore, Style
from pprint import pprint


def print_red(text):
    print(Fore.RED + text)
    print(Style.RESET_ALL)


def print_green(text):
    print(Fore.GREEN + text)
    print(Style.RESET_ALL)


def print_before_and_after(main_chracter, title=""):
    def decorator(func):
        def wrapper(*args, **kwargs):
            print_green(main_chracter * 10)
            if title:
                print_green(title)
            result = func(*args, **kwargs)
            print_green(main_chracter * 10)
            return result

        return wrapper

    return decorator


######################


def get_sheet(filepath, sheet_name) -> pd.DataFrame | None:
    if sheet_name == "Opening File":
        return pd.read_excel(
            filepath,
            engine="openpyxl",
            sheet_name=sheet_name,
            nrows=None,
            skiprows=16,
        )
    return pd.read_excel(filepath, nrows=None, sheet_name=sheet_name, engine="openpyxl")


def clear_sheet(sheet_name, filepath):
    workbook = openpyxl.load_workbook(filepath)
    sheet_workbook = workbook[sheet_name]
    first_row = True
    for row in sheet_workbook.iter_rows():
        if first_row:
            first_row = False
            continue

        for cell in row:
            cell.value = None

    workbook.save(filepath)
    workbook.close()


def write_into_sheet(sheet_name, df, filepath):
    try:
        try:
            # writer = pd.ExcelWriter(filepath, engine="openpyxl", mode="w")
            writer = pd.ExcelWriter(
                filepath, engine="openpyxl", mode="a", if_sheet_exists="overlay"
            )
        except zipfile.BadZipFile as e:
            print_red(f"File integrity issue: {e}")
            return
        except PermissionError as e:
            print_red(f"Permission denied: {e}")
            return
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.book.save(filepath)
        # this line have make problem for all file
        # writer.close()

    except Exception as e:
        print(f"error for writing into file in sheet {sheet_name}, {e}")


# @print_before_and_after("~", "into handel query function: ")
def process_excel_queries(
    filepath, sheet_name, queries, log=False
) -> pd.DataFrame | None:
    """Processes a series of Excel query-like operations on a sheet.

    Args:
        filepath (str): The path to the Excel workbook as a string.
        sheet_name (str): The name of the sheet containing the data as a string.
        queries (list): A list of dictionaries representing query operations.
            Each dictionary should have keys:
                - "operation" (str): The operation to perform (e.g., "read", "change_type", "remove_columns", "filter_rows").
                - "arguments" (list, optional): A list of arguments specific to the operation.

    Returns:
        pandas.DataFrame: A pandas DataFrame containing the processed data, or None if errors occur.
    """
    # NOTE : maybe for date column must date datetiem field

    # NOTE: Read data from the target sheet
    # NOTE it is more than it need but for dont take error set it
    df = get_sheet(filepath, sheet_name)
    operation = ""
    try:
        # Process queries sequentially
        # print_red("befor set query filters")
        for query in queries:
            operation = query["operation"].lower()
            arguments = query.get("arguments", [])

            if operation == "read":
                # Read data if not already read
                sh_name = arguments["sheet_name"]
                # if df is None:
                df = get_sheet(filepath, sh_name)
            # NOTE: this filter isn't set for all queries Because it make some issue
            # NOTE: chage header if needed
            elif operation == "change_type":
                continue
                # print_green(df["Date Opened"])
                # if haeder and arguments of change type are note same it will add arguments into header
                try:
                    if list(df.columns) != [i for i, _ in arguments]:
                        # print_red("header of list and sheet not same ")
                        # Update the header row
                        update_header_list = []
                        update_header_list.extend(df.columns)
                        for header, _ in arguments:
                            if header not in df.columns:
                                update_header_list.append(header)

                        # pprint(update_header_list)
                        workbook = openpyxl.load_workbook(filepath)
                        # clear_sheet(workbook, sheet_name)
                        write_header(
                            workbook=workbook,
                            target_sh_name=sheet_name,
                            source_sh="",
                            start_header_row=1,
                            header_list=update_header_list,
                        )
                        workbook.save(filepath)
                except Exception as e:
                    print_red(f"error in change number of header")
                    # df = get_sheet(filepath, sheet_name)
                    # print_green(df)
                # Set data types for specific columns
                try:
                    for col_name, col_type in arguments:
                        df[col_name] = df[col_name].astype(col_type)
                except Exception as e:
                    print(f"error in chage type of columns {e}")

            elif operation == "remove_columns":
                for column in arguments:
                    try:
                        df = df.drop(column, axis=1)
                    except Exception as e:
                        if log:
                            print_red(
                                f"can't delete columns of sheet '{sheet_name}': {e}"
                            )

            # NOTE: rows that have "\n" will be select
            elif operation == "combine_sheets":
                data_frames = []
                # just move rows that have data
                for sht_name in arguments:
                    sh_df = get_sheet(filepath, sht_name)
                    # filter sheet data with have any data cell
                    sh_df = sh_df[sh_df.iloc[:, -len(df.columns) :].any(axis=1)]
                    if sht_name == "Magistrates":
                        sh_df = sh_df[sh_df.iloc[:, 0].notna()]
                    data_frames.append(sh_df)
                if log:
                    print_green("dataframes in combine_sheets:")
                    if log:
                        for i in data_frames:
                            print(i)
                df = pd.concat(data_frames, ignore_index=True)

            # NOTE: must check all conditions
            # BUG: check have yes conditions and Variable that have space in it
            elif operation == "filter_rows":
                filter_condition = arguments[0]

                ###########
                # handel sheet that have space
                if filter_condition == "'Type of Offence' == 'Either Way'":
                    df = df[df["Type of Offence"] == "Either Way"]
                elif filter_condition == "'Type of Offence' == 'Indictable'":
                    df = df[df["Type of Offence"] == "Indictable"]
                ###########
                else:
                    df = df.query(filter_condition)
                    # df = df[df["Court"] == "Police Station"]  # it work for police Station

            elif operation == "select_columns":
                continue
                df = df[arguments]
            else:
                print(f"Warning: Unsupported operation '{operation}'.")

            if log:
                print_green(f"df in operation '{operation}'")
                try:
                    print_green(f"\t{df.shape[0]} row")
                except Exception as e:
                    print()
                    print_red("error for get dataframe")
                    print_red(e)
                    pass

        return df

    except FileNotFoundError:
        print(f"Error: File '{filepath}' not found.")
        return None
    except Exception as e:
        print(f"Error processing queries: {e}")
        print(f"in sheet '{sheet_name}' and operation '{operation}' ")
        return None


@print_before_and_after("###")
def main(filepath, can_write=True):
    for item in queries:
        sheet_name = item["item_name"]

        # clear sheet befor write into it
        # NOTE: if clear "police Station" sheet befor write into it, bail sheet will empty
        if sheet_name != "Police Station":
            clear_sheet(sheet_name, filepath)

        queries_list = []
        for query in item["quires"]:
            queries_list.append(query)

        df = process_excel_queries(
            filepath,
            sheet_name,
            queries_list,
        )
        try:
            print(f"sheet '{sheet_name}' have {df.shape[0]} row")
        except:
            print(f"error for show df in sheet {sheet_name}")

        # write queries of sheet
        try:
            if can_write:
                write_into_sheet(sheet_name, df, filepath)
                print(f"'{sheet_name}' saved")
        except Exception as e:
            print(e)


if __name__ == "__main__":
    from query_config import *

    main(filepath, can_write)
else:
    from utils.query_list import queries

    # from utils.btn import write_header
